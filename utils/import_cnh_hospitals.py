"""Import the Catálogo Nacional de Hospitales (CNH) 2025 for the QoL slice.

Downloads the Ministerio de Sanidad's CNH 2025 Excel, keeps the hospitals in
the five provinces the owner's searches cover (A Coruña, Lugo, Ourense,
Pontevedra, Asturias), and writes `data/hospitals_cnh.json` with the raw CNH
fields plus a *descriptive* grouping agreed with the owner on 2026-08-13.

    python -m utils.import_cnh_hospitals --out data/hospitals_cnh.json
    python -m utils.import_cnh_hospitals --out data/hospitals_cnh.json --skip-geocode
    python -m utils.import_cnh_hospitals --out data/hospitals_cnh.json --geocode-only
    python -m utils.import_cnh_hospitals --out ... --xlsx /path/to/CNH_2025.xlsx

Column layout — measured against the real `CNH_2025.xlsx` on 2026-08-14, not
assumed from older editions:

- Sheet ``DIRECTORIO DE HOSPITALES`` (one row per hospital): ``CCN``,
  ``CODCNH``, ``Nombre Centro``, ``Dirección``, ``Teléfono``,
  ``Cód. Municipio``, ``Municipio``, ``Cód. Provincia``, ``Provincia``,
  ``Cód. CCAA``, ``CCAA``, ``Código Postal``, ``CAMAS``,
  ``Cód. Clase de Centro``, ``Clase de Centro``, ``Cód. Dep. Funcional``,
  ``Dependencia Funcional``, ``Forma parte Complejo``, ``CODIDCOM``,
  ``Nombre del Complejo``, ``ALTA``, ``Email``.
- Sheet ``ESTRUCTURA FUNCIONAL``: ``CODCNH``, ``AD`` (acreditación docente,
  S/N), ``Forma parte Complejo``, ``CODIDCOM``, ``Es Comlejo`` (sic — the
  header really is misspelled in the file), and the alta-tecnología equipment
  unit counts ``TAC RMN GAM HEM ASD LIT BCO ALI SPECT PET MAMO DO DIAL``.
  A cell holding ``DC`` means the datum is declared on the hospital's complex
  row (``Es Comlejo`` = S) instead of the member row.

The 2025 workbook carries no literal ``FINALIDAD ASISTENCIAL`` column; its
equivalent is ``Clase de Centro``, where code C11 = "Hospitales Generales" is
the finalidad-GENERAL population. C13 (media y larga estancia) and C14 (salud
mental) are the monographics the owner asked to drop; C12 (especializados) and
C190 (otros con internamiento) are not general/acute either. If a province
were left with nothing after that filter, its non-general hospitals are kept
so the province is not silently empty — the entry's raw fields show what it is.

``high_tech_count`` counts *distinct* equipment types with at least one unit
among TAC, RMN, GAM, HEM, ASD, LIT, BCO, ALI, SPECT, PET. MAMO (mammograph),
DO (bone densitometer) and DIAL (dialysis stations) are recorded in the
``equipment`` breakdown but deliberately not counted: they are routine
capability, and DIAL alone would push every district hospital with a dialysis
unit over the teaching_high_tech threshold. The per-type units are stored so a
consumer can show the derivation.

Grouping — descriptive, NOT an official CNH classification (owner, 2026-08-13):

- ``teaching_high_tech``: teaching-accredited AND high_tech_count >= 3
- ``general_acute``: everything else with beds >= 100
- ``limited_recorded_capability``: the rest

Coordinates come from the app's own geocoding helper
(`utils.geocoding.GeocodingService.geocode_address`) — a few dozen one-off
calls. `--skip-geocode` writes lat/lon null with ``"geocode": "skipped"`` (for
CI and tests); a failed or implausible geocode stays null with
``"geocode": "failed"`` — never a guessed coordinate. `--geocode-only` re-reads
an existing output file and fills in only the entries that are not ``ok``,
which lets the parse run where openpyxl lives and the geocode run where the
Google key lives (the app container).
"""

import argparse
import json
import logging
import os
import re
import sys
import tempfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

CNH_PAGE_URL = (
    "https://www.sanidad.gob.es/estadEstudios/estadisticas/sisInfSanSNS/"
    "ofertaRecursos/hospitales/home.htm"
)
CNH_XLSX_URL = (
    "https://www.sanidad.gob.es/estadEstudios/estadisticas/sisInfSanSNS/"
    "ofertaRecursos/hospitales/docs/CNH_2025.xlsx"
)
# The real file is ~0.5 MB; anything past this is not the catalogue.
MAX_XLSX_BYTES = 20 * 1024 * 1024

DIRECTORIO_SHEET = "DIRECTORIO DE HOSPITALES"
ESTRUCTURA_SHEET = "ESTRUCTURA FUNCIONAL"

# Cód. Provincia -> name as this project spells it. The sheet's own
# `Provincia` column agrees, so it is a filter key, not a rename.
TARGET_PROVINCES = {
    "15": "A Coruña",
    "27": "Lugo",
    "32": "Ourense",
    "33": "Asturias",
    "36": "Pontevedra",
}

# Clase de Centro C11 = "Hospitales Generales" — the finalidad-GENERAL rows.
GENERAL_CLASS_CODE = "C11"

# Counted toward high_tech_count vs recorded-only (see module docstring).
HIGH_TECH_COUNTED = (
    "TAC",
    "RMN",
    "GAM",
    "HEM",
    "ASD",
    "LIT",
    "BCO",
    "ALI",
    "SPECT",
    "PET",
)
EQUIPMENT_RECORDED_ONLY = ("MAMO", "DO", "DIAL")
EQUIPMENT_COLUMNS = HIGH_TECH_COUNTED + EQUIPMENT_RECORDED_ONLY

GROUP_TEACHING_HIGH_TECH = "teaching_high_tech"
GROUP_GENERAL_ACUTE = "general_acute"
GROUP_LIMITED = "limited_recorded_capability"

TEACHING_HIGH_TECH_MIN_ITEMS = 3
GENERAL_ACUTE_MIN_BEDS = 100

# Plausibility box for the five provinces. A geocode that lands outside it
# matched some other place in the world; storing it would be a guessed
# coordinate wearing an "ok" badge.
REGION_LAT_RANGE = (41.5, 44.2)
REGION_LON_RANGE = (-9.5, -4.3)

# "Coruña, A" / "Barco de Valdeorras, O" -> natural order for display and for
# the geocoder. Only the article forms Spanish/Galician municipality names use.
_INVERTED_ARTICLE = re.compile(
    r"^(?P<name>.+),\s*(?P<article>A|O|As|Os|El|La|Los|Las|L')$"
)

DIRECTORIO_REQUIRED = (
    "CODCNH",
    "Nombre Centro",
    "Municipio",
    "Cód. Provincia",
    "Provincia",
    "CAMAS",
    "Cód. Clase de Centro",
    "Clase de Centro",
    "Dependencia Funcional",
)
# "Es Comlejo" is the file's own misspelling of "Es Complejo".
ESTRUCTURA_REQUIRED = ("CODCNH", "AD", "CODIDCOM", "Es Comlejo") + EQUIPMENT_COLUMNS


def classify_grouping(
    teaching: Optional[bool], high_tech_count: int, beds: Optional[int]
) -> str:
    """The descriptive grouping agreed with the owner on 2026-08-13.

    An unknown teaching flag or bed count never *promotes* a hospital: it
    simply cannot satisfy the clause that needs the missing datum.
    """
    if teaching and high_tech_count >= TEACHING_HIGH_TECH_MIN_ITEMS:
        return GROUP_TEACHING_HIGH_TECH
    if beds is not None and beds >= GENERAL_ACUTE_MIN_BEDS:
        return GROUP_GENERAL_ACUTE
    return GROUP_LIMITED


def normalize_municipality(name: str) -> str:
    """Undo the catalogue's inverted-article form: "Coruña, A" -> "A Coruña"."""
    match = _INVERTED_ARTICLE.match(name.strip())
    if not match:
        return name.strip()
    article = match.group("article")
    base = match.group("name").strip()
    if article == "L'":
        return f"L'{base}"
    return f"{article} {base}"


def _header_map(
    header_row: Tuple[Any, ...], required: Tuple[str, ...], sheet: str
) -> Dict[str, int]:
    headers = {
        str(cell).strip(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None
    }
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(
            f"Sheet {sheet!r} is missing expected columns: {missing} — the CNH layout changed; re-measure before trusting this parser."
        )
    return headers


def _as_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "DC":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _resolve_dc(value: Any, complex_row: Optional[Tuple[Any, ...]], idx: int) -> Any:
    """A `DC` cell defers to the hospital's complex row; unresolvable stays DC."""
    if str(value).strip().upper() != "DC":
        return value
    if complex_row is None:
        return value
    return complex_row[idx]


def _teaching_from_ad(value: Any) -> Optional[bool]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if text == "S":
        return True
    if text == "N":
        return False
    # `DC` that could not be resolved, or anything unexpected: unknown, not False.
    return None


# The download cap bounds the *compressed* file; a crafted workbook could
# still expand far past it. The declared decompressed sizes are checked
# before openpyxl touches the archive (diff review, 2026-08-14).
MAX_DECOMPRESSED_BYTES = 100 * 1024 * 1024


def parse_workbook(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parse the CNH workbook into hospital entries for the target provinces."""
    import openpyxl  # heavy and only needed for the parse phase

    import zipfile

    with zipfile.ZipFile(xlsx_path) as archive:
        declared = sum(info.file_size for info in archive.infolist())
    if declared > MAX_DECOMPRESSED_BYTES:
        raise ValueError(
            f"Workbook declares {declared} decompressed bytes "
            f"(cap {MAX_DECOMPRESSED_BYTES}) — refusing to parse"
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet in (DIRECTORIO_SHEET, ESTRUCTURA_SHEET):
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Workbook has no sheet {sheet!r} — sheets present: {wb.sheetnames}"
            )

    est_rows = list(wb[ESTRUCTURA_SHEET].iter_rows(values_only=True))
    est_cols = _header_map(est_rows[0], ESTRUCTURA_REQUIRED, ESTRUCTURA_SHEET)
    est_by_code: Dict[str, Tuple[Any, ...]] = {}
    complex_by_code: Dict[str, Tuple[Any, ...]] = {}
    for row in est_rows[1:]:
        code = row[est_cols["CODCNH"]]
        if code is None:
            continue
        code = str(code).strip()
        if str(row[est_cols["Es Comlejo"]] or "").strip().upper() == "S":
            complex_by_code[code] = row
        else:
            est_by_code[code] = row

    dir_rows = list(wb[DIRECTORIO_SHEET].iter_rows(values_only=True))
    dir_cols = _header_map(dir_rows[0], DIRECTORIO_REQUIRED, DIRECTORIO_SHEET)

    by_province: Dict[str, List[Dict[str, Any]]] = {
        code: [] for code in TARGET_PROVINCES
    }
    for row in dir_rows[1:]:
        prov_code = str(row[dir_cols["Cód. Provincia"]] or "").strip()
        if prov_code not in TARGET_PROVINCES:
            continue
        cnh_code = str(row[dir_cols["CODCNH"]] or "").strip()
        est = est_by_code.get(cnh_code)
        complex_row = None
        if est is not None:
            codidcom = est[est_cols["CODIDCOM"]]
            if codidcom is not None:
                complex_row = complex_by_code.get(str(codidcom).strip())

        teaching: Optional[bool] = None
        equipment: Dict[str, int] = {}
        high_tech_count = 0
        if est is None:
            logger.warning(
                "Hospital %s (%s) has no ESTRUCTURA FUNCIONAL row: teaching and equipment unknown",
                cnh_code,
                row[dir_cols["Nombre Centro"]],
            )
        else:
            ad = _resolve_dc(est[est_cols["AD"]], complex_row, est_cols["AD"])
            teaching = _teaching_from_ad(ad)
            for col in EQUIPMENT_COLUMNS:
                idx = est_cols[col]
                units = _as_int(_resolve_dc(est[idx], complex_row, idx))
                if units is not None and units > 0:
                    equipment[col] = units
                    if col in HIGH_TECH_COUNTED:
                        high_tech_count += 1

        beds = _as_int(row[dir_cols["CAMAS"]])
        clase_cod = str(row[dir_cols["Cód. Clase de Centro"]] or "").strip()
        clase_text = str(row[dir_cols["Clase de Centro"]] or "").strip()
        entry = {
            "cnh_code": cnh_code,
            "name": str(row[dir_cols["Nombre Centro"]] or "").strip(),
            "municipality": normalize_municipality(
                str(row[dir_cols["Municipio"]] or "")
            ),
            "province": str(row[dir_cols["Provincia"]] or "").strip(),
            "beds": beds,
            "teaching": teaching,
            "high_tech_count": high_tech_count,
            "equipment": equipment,
            "dependencia": str(row[dir_cols["Dependencia Funcional"]] or "").strip(),
            "finalidad": "GENERAL" if clase_cod == GENERAL_CLASS_CODE else clase_text,
            "clase_centro_cod": clase_cod,
            "clase_centro": clase_text,
            "grouping": classify_grouping(teaching, high_tech_count, beds),
            "lat": None,
            "lon": None,
            "geocode": "skipped",
        }
        by_province[prov_code].append(entry)

    hospitals: List[Dict[str, Any]] = []
    for prov_code, entries in by_province.items():
        general = [e for e in entries if e["clase_centro_cod"] == GENERAL_CLASS_CODE]
        if general:
            kept = general
        else:
            # A province with no general hospital keeps what it has rather than
            # vanishing; the raw clase/finalidad fields say what these are.
            kept = entries
            if entries:
                logger.warning(
                    "Province %s has no %s (general) hospital; keeping its %d non-general entries",
                    TARGET_PROVINCES[prov_code],
                    GENERAL_CLASS_CODE,
                    len(entries),
                )
        hospitals.extend(kept)

    hospitals.sort(key=lambda e: (e["province"], e["municipality"], e["name"]))
    return hospitals


def download_xlsx(dest_path: str) -> None:
    """Fetch the CNH Excel from the Ministerio de Sanidad, bounded and verified."""
    import requests

    from utils.http import HTTP_USER_AGENT, request_with_retries

    response = request_with_retries(
        requests.get,
        CNH_XLSX_URL,
        headers={"User-Agent": HTTP_USER_AGENT},
        timeout=60,
        stream=True,
        logger=logger,
    )
    try:
        if response.status_code != 200:
            raise RuntimeError(
                f"CNH Excel download failed: HTTP {response.status_code} from {CNH_XLSX_URL}"
            )
        size = 0
        with open(dest_path, "wb") as handle:
            for chunk in response.iter_content(chunk_size=65536):
                size += len(chunk)
                if size > MAX_XLSX_BYTES:
                    raise RuntimeError(
                        f"CNH Excel exceeds {MAX_XLSX_BYTES} bytes — refusing to parse it"
                    )
                handle.write(chunk)
    finally:
        response.close()
    with open(dest_path, "rb") as handle:
        magic = handle.read(4)
    if not magic.startswith(b"PK"):
        raise RuntimeError(
            f"Downloaded file from {CNH_XLSX_URL} is not an .xlsx (bad magic {magic!r}) — "
            "the ministry may have moved the catalogue; find the new URL, do not scrape the PDF."
        )


def geocode_hospitals(hospitals: List[Dict[str, Any]]) -> Tuple[int, int]:
    """Fill lat/lon via the app's Google geocoding helper. Returns (ok, failed).

    Only entries not already ``ok`` are looked up, so a `--geocode-only` rerun
    retries failures without re-spending calls on successes.
    """
    from utils.geocoding import GeocodingService

    service = GeocodingService()
    ok = failed = 0
    for entry in hospitals:
        if entry.get("geocode") == "ok" and entry.get("lat") is not None:
            ok += 1
            continue
        address = (
            f"{entry['name']}, {entry['municipality']}, {entry['province']}, Spain"
        )
        result = service.geocode_address(address)
        lat = result.get("lat") if result else None
        lon = result.get("lng") if result else None
        plausible = (
            isinstance(lat, (int, float))
            and isinstance(lon, (int, float))
            and REGION_LAT_RANGE[0] <= lat <= REGION_LAT_RANGE[1]
            and REGION_LON_RANGE[0] <= lon <= REGION_LON_RANGE[1]
        )
        if plausible:
            entry["lat"] = round(float(lat), 6)
            entry["lon"] = round(float(lon), 6)
            entry["geocode"] = "ok"
            accuracy = result.get("accuracy")
            if accuracy:
                entry["geocode_accuracy"] = accuracy
            ok += 1
        else:
            if result is not None:
                logger.warning(
                    "Geocode for %r landed outside the region box (%r, %r): recorded as failed",
                    address,
                    lat,
                    lon,
                )
            entry["lat"] = None
            entry["lon"] = None
            entry["geocode"] = "failed"
            entry.pop("geocode_accuracy", None)
            failed += 1
    return ok, failed


def build_document(hospitals: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": f"CNH 2025, Ministerio de Sanidad ({CNH_XLSX_URL})",
        "distance_basis_note": "coords via Google Geocoding, one-off",
        "grouping_note": (
            "Descriptive grouping, not an official CNH classification: "
            "teaching_high_tech = teaching AND high_tech_count >= "
            f"{TEACHING_HIGH_TECH_MIN_ITEMS}; general_acute = beds >= "
            f"{GENERAL_ACUTE_MIN_BEDS}; limited_recorded_capability = the rest."
        ),
        "high_tech_note": (
            "high_tech_count = distinct equipment types with >= 1 unit among "
            + "/".join(HIGH_TECH_COUNTED)
            + " (ESTRUCTURA FUNCIONAL sheet); "
            + "/".join(EQUIPMENT_RECORDED_ONLY)
            + " are recorded in `equipment` but not counted."
        ),
        "hospitals": hospitals,
    }


def write_document(document: Dict[str, Any], out_path: str) -> None:
    """Owner-only temp file in the target directory, fsync, atomic rename."""
    out_path = os.path.abspath(out_path)
    directory = os.path.dirname(out_path)
    os.makedirs(directory, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        dir=directory, prefix=".hospitals_cnh.", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(document, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.chmod(tmp_path, 0o644)
        os.replace(tmp_path, out_path)
    except BaseException:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def report(document: Dict[str, Any]) -> str:
    hospitals = document["hospitals"]
    by_province: Dict[str, int] = {}
    by_grouping: Dict[str, int] = {}
    geocode: Dict[str, int] = {}
    for entry in hospitals:
        by_province[entry["province"]] = by_province.get(entry["province"], 0) + 1
        by_grouping[entry["grouping"]] = by_grouping.get(entry["grouping"], 0) + 1
        geocode[entry["geocode"]] = geocode.get(entry["geocode"], 0) + 1
    lines = [f"hospitals: {len(hospitals)}"]
    lines.append(
        "per province: " + ", ".join(f"{k}={v}" for k, v in sorted(by_province.items()))
    )
    lines.append(
        "per grouping: " + ", ".join(f"{k}={v}" for k, v in sorted(by_grouping.items()))
    )
    lines.append(
        "geocode: " + ", ".join(f"{k}={v}" for k, v in sorted(geocode.items()))
    )
    return "\n".join(lines)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import the CNH 2025 hospital catalogue for the target provinces"
    )
    parser.add_argument(
        "--out", required=True, help="output JSON path (data/hospitals_cnh.json)"
    )
    parser.add_argument(
        "--xlsx",
        help="use a local CNH_2025.xlsx instead of downloading it",
    )
    parser.add_argument(
        "--skip-geocode",
        action="store_true",
        help="write lat/lon null with geocode=skipped (no Google calls)",
    )
    parser.add_argument(
        "--geocode-only",
        action="store_true",
        help="re-read --out and geocode only its non-ok entries (no Excel parse)",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO, format="%(levelname)s %(name)s: %(message)s"
    )

    if args.geocode_only and args.skip_geocode:
        parser.error("--geocode-only and --skip-geocode contradict each other")

    if args.geocode_only:
        with open(args.out, encoding="utf-8") as handle:
            document = json.load(handle)
        hospitals = document["hospitals"]
    else:
        if args.xlsx:
            hospitals = parse_workbook(args.xlsx)
        else:
            with tempfile.TemporaryDirectory() as tmpdir:
                xlsx_path = os.path.join(tmpdir, "CNH_2025.xlsx")
                download_xlsx(xlsx_path)
                hospitals = parse_workbook(xlsx_path)
        document = build_document(hospitals)

    if not hospitals:
        raise RuntimeError(
            "Parsed zero hospitals for the target provinces — refusing to write an empty catalogue"
        )

    if not args.skip_geocode:
        ok, failed = geocode_hospitals(hospitals)
        logger.info("geocoded: %d ok, %d failed", ok, failed)
        document["generated_at"] = datetime.now(timezone.utc).isoformat()

    write_document(document, args.out)
    print(report(document))
    print(f"written: {os.path.abspath(args.out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
