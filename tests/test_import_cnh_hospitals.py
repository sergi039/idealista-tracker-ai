"""utils/import_cnh_hospitals.py — grouping rule, parser, and honest coords.

The worksheet fixture reproduces the real CNH_2025.xlsx layout as measured on
2026-08-14: sheet "DIRECTORIO DE HOSPITALES" (CCN, CODCNH, Nombre Centro,
Dirección, Teléfono, Cód. Municipio, Municipio, Cód. Provincia, Provincia,
Cód. CCAA, CCAA, Código Postal, CAMAS, Cód. Clase de Centro, Clase de Centro,
Cód. Dep. Funcional, Dependencia Funcional, Forma parte Complejo, CODIDCOM,
Nombre del Complejo, ALTA, Email) and sheet "ESTRUCTURA FUNCIONAL" (CCN,
CODCNH, Nombre Centro, Cód. Provincia, Provincia, Cód. CCAA, CCAA, AD,
Cod Concierto, Concierto, Forma parte Complejo, CODIDCOM, Nombre del Complejo,
"Es Comlejo" — the file's own misspelling — ALTA, then the equipment columns
TAC RMN GAM HEM ASD LIT BCO ALI SPECT PET MAMO DO DIAL). `DC` in a member
row's cell defers to the complex row.
"""

import json

import openpyxl
import pytest

from utils.import_cnh_hospitals import (
    GROUP_GENERAL_ACUTE,
    GROUP_LIMITED,
    GROUP_TEACHING_HIGH_TECH,
    classify_grouping,
    main,
    normalize_municipality,
    parse_workbook,
)

DIRECTORIO_HEADER = [
    "CCN",
    "CODCNH",
    "Nombre Centro",
    "Dirección",
    "Teléfono",
    "Cód. Municipio",
    "Municipio",
    "Cód. Provincia",
    "Provincia",
    "Cód. CCAA",
    "CCAA",
    "Código Postal",
    "CAMAS",
    "Cód. Clase de Centro",
    "Clase de Centro",
    "Cód. Dep. Funcional",
    "Dependencia Funcional",
    "Forma parte Complejo",
    "CODIDCOM",
    "Nombre del Complejo",
    "ALTA",
    "Email",
]

EQUIPMENT = [
    "TAC",
    "RMN",
    "GAM",
    "HEM",
    "ASD",
    "LIT",
    "BCO",
    "ALI",
    "SPECT",
    "PET",
    "MAMO",
    "DO",
    "DIAL",
]

ESTRUCTURA_HEADER = [
    "CCN",
    "CODCNH",
    "Nombre Centro",
    "Cód. Provincia",
    "Provincia",
    "Cód. CCAA",
    "CCAA",
    "AD",
    "Cod Concierto",
    "Concierto",
    "Forma parte Complejo",
    "CODIDCOM",
    "Nombre del Complejo",
    "Es Comlejo",
    "ALTA",
] + EQUIPMENT


def dir_row(
    codcnh,
    name,
    municipality,
    prov_cod,
    provincia,
    camas,
    clase_cod,
    clase,
    dependencia="Privados",
    forma_parte="N",
    codidcom=None,
):
    return [
        None,
        codcnh,
        name,
        "Calle Falsa 1",
        None,
        None,
        municipality,
        prov_cod,
        provincia,
        None,
        None,
        None,
        camas,
        clase_cod,
        clase,
        None,
        dependencia,
        forma_parte,
        codidcom,
        None,
        "N",
        None,
    ]


def est_row(codcnh, ad, equipment=None, es_complejo="N", codidcom=None):
    if equipment == "DC":
        values = ["DC"] * len(EQUIPMENT)
    else:
        equipment = equipment or {}
        values = [str(equipment.get(col, 0)) for col in EQUIPMENT]
    return [
        None,
        codcnh,
        None,
        None,
        None,
        None,
        None,
        ad,
        None,
        None,
        "S" if codidcom else "N",
        codidcom,
        None,
        es_complejo,
        "N",
    ] + values


def write_fixture(path, directorio_rows, estructura_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIRECTORIO DE HOSPITALES"
    ws.append(DIRECTORIO_HEADER)
    for row in directorio_rows:
        ws.append(row)
    est = wb.create_sheet("ESTRUCTURA FUNCIONAL")
    est.append(ESTRUCTURA_HEADER)
    for row in estructura_rows:
        est.append(row)
    wb.save(path)
    return str(path)


@pytest.fixture
def fixture_xlsx(tmp_path):
    """A small catalogue exercising every parser rule at once."""
    directorio = [
        # Teaching + 3 counted high-tech types -> teaching_high_tech. DIAL is
        # recorded but must not be the third counted type.
        dir_row(
            "330001",
            "Hospital Universitario Test",
            "Oviedo",
            "33",
            "Asturias",
            "900",
            "C11",
            "Hospitales Generales",
            dependencia="Servicios e Institutos de Salud de Las Comunidades Autónomas",
        ),
        # Teaching but only ONE counted type (TAC; MAMO/DIAL don't count):
        # falls through to the beds rule -> general_acute.
        dir_row(
            "330002",
            "Hospital Comarcal Test",
            "Coaña",
            "33",
            "Asturias",
            "116",
            "C11",
            "Hospitales Generales",
        ),
        # Small, nothing recorded -> limited_recorded_capability; the
        # inverted-article municipality is normalized.
        dir_row(
            "150001",
            "Clinica Pequena Test",
            "Coruña, A",
            "15",
            "A Coruña",
            "40",
            "C11",
            "Hospitales Generales",
        ),
        # Salud mental in a province that has a general hospital -> dropped.
        dir_row(
            "330003",
            "Psiquiatrico Test",
            "Gijón",
            "33",
            "Asturias",
            "60",
            "C14",
            "Hospitales de salud mental y tratamiento de toxicomanías",
        ),
        # Non-target province -> dropped.
        dir_row(
            "280001",
            "Hospital Madrid Test",
            "Madrid",
            "28",
            "Madrid",
            "500",
            "C11",
            "Hospitales Generales",
        ),
        # Complex member: its AD and equipment cells hold DC and defer to the
        # complex row 270999 -> resolved to teaching + 3 types.
        dir_row(
            "270001",
            "Hospital Miembro Complexo",
            "Lugo",
            "27",
            "Lugo",
            "300",
            "C11",
            "Hospitales Generales",
            forma_parte="S",
            codidcom="270999",
        ),
        # Ourense holds ONLY a media/larga estancia hospital: the per-province
        # fallback keeps it, with its real (non-GENERAL) finalidad on show.
        dir_row(
            "320001",
            "Hospital Larga Estancia Test",
            "Ourense",
            "32",
            "Ourense",
            "80",
            "C13",
            "Hospitales de media y larga estancia",
        ),
    ]
    estructura = [
        est_row("330001", "S", {"TAC": 2, "RMN": 1, "ALI": 1, "DIAL": 40}),
        est_row("330002", "S", {"TAC": 1, "MAMO": 1, "DIAL": 8}),
        est_row("150001", "N"),
        est_row("330003", "N"),
        est_row("280001", "S", {"TAC": 5, "RMN": 3, "PET": 1}),
        est_row("270001", "DC", "DC", codidcom="270999"),
        est_row("270999", "S", {"TAC": 1, "RMN": 1, "HEM": 1}, es_complejo="S"),
        est_row("320001", "N"),
    ]
    return write_fixture(tmp_path / "cnh_fixture.xlsx", directorio, estructura)


class TestGroupingBoundaries:
    def test_teaching_and_three_high_tech_is_teaching_high_tech(self):
        assert classify_grouping(True, 3, 50) == GROUP_TEACHING_HIGH_TECH

    def test_teaching_with_two_high_tech_falls_to_beds_rule(self):
        assert classify_grouping(True, 2, 500) == GROUP_GENERAL_ACUTE

    def test_high_tech_without_teaching_is_not_teaching_high_tech(self):
        assert classify_grouping(False, 10, 500) == GROUP_GENERAL_ACUTE

    def test_unknown_teaching_never_promotes(self):
        assert classify_grouping(None, 10, 500) == GROUP_GENERAL_ACUTE

    def test_beds_boundary_100_is_general_acute(self):
        assert classify_grouping(False, 0, 100) == GROUP_GENERAL_ACUTE

    def test_beds_99_is_limited(self):
        assert classify_grouping(False, 0, 99) == GROUP_LIMITED

    def test_unknown_beds_never_promotes(self):
        assert classify_grouping(False, 0, None) == GROUP_LIMITED

    def test_teaching_high_tech_needs_no_beds(self):
        assert classify_grouping(True, 3, None) == GROUP_TEACHING_HIGH_TECH


class TestParser:
    def test_parses_filters_and_groups(self, fixture_xlsx):
        hospitals = parse_workbook(fixture_xlsx)
        by_code = {h["cnh_code"]: h for h in hospitals}

        # Dropped: non-target province, and C14 in a province holding a C11.
        assert "280001" not in by_code
        assert "330003" not in by_code
        assert set(by_code) == {"330001", "330002", "150001", "270001", "320001"}

        uni = by_code["330001"]
        assert uni["grouping"] == GROUP_TEACHING_HIGH_TECH
        assert uni["teaching"] is True
        assert uni["high_tech_count"] == 3
        assert uni["equipment"] == {"TAC": 2, "RMN": 1, "ALI": 1, "DIAL": 40}
        assert uni["beds"] == 900
        assert uni["finalidad"] == "GENERAL"
        assert uni["clase_centro_cod"] == "C11"

        comarcal = by_code["330002"]
        # MAMO and DIAL are recorded but not counted: one counted type only,
        # so teaching alone does not make this teaching_high_tech.
        assert comarcal["high_tech_count"] == 1
        assert comarcal["equipment"] == {"TAC": 1, "MAMO": 1, "DIAL": 8}
        assert comarcal["grouping"] == GROUP_GENERAL_ACUTE

        small = by_code["150001"]
        assert small["grouping"] == GROUP_LIMITED
        assert small["municipality"] == "A Coruña"

    def test_dc_cells_resolve_through_the_complex_row(self, fixture_xlsx):
        by_code = {h["cnh_code"]: h for h in parse_workbook(fixture_xlsx)}
        member = by_code["270001"]
        assert member["teaching"] is True
        assert member["high_tech_count"] == 3
        assert member["equipment"] == {"TAC": 1, "RMN": 1, "HEM": 1}
        assert member["grouping"] == GROUP_TEACHING_HIGH_TECH
        # The complex aggregate row itself is not a hospital entry.
        assert "270999" not in by_code

    def test_province_with_no_general_hospital_keeps_what_it_has(self, fixture_xlsx):
        by_code = {h["cnh_code"]: h for h in parse_workbook(fixture_xlsx)}
        fallback = by_code["320001"]
        assert fallback["province"] == "Ourense"
        # Its real finalidad stays on show — never relabeled GENERAL.
        assert fallback["finalidad"] == "Hospitales de media y larga estancia"
        assert fallback["clase_centro_cod"] == "C13"

    def test_missing_expected_column_fails_loudly(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DIRECTORIO DE HOSPITALES"
        ws.append(["CODCNH", "Nombre Centro"])  # layout drifted
        est = wb.create_sheet("ESTRUCTURA FUNCIONAL")
        est.append(ESTRUCTURA_HEADER)
        path = tmp_path / "drifted.xlsx"
        wb.save(path)
        with pytest.raises(ValueError, match="missing expected columns"):
            parse_workbook(str(path))


class TestNoInventedCoords:
    def test_skip_geocode_writes_null_coords_and_never_geocodes(
        self, fixture_xlsx, tmp_path, monkeypatch
    ):
        def _boom(*_args, **_kwargs):
            raise AssertionError("--skip-geocode must not geocode")

        monkeypatch.setattr("utils.import_cnh_hospitals.geocode_hospitals", _boom)
        out = tmp_path / "hospitals.json"
        main(["--out", str(out), "--xlsx", fixture_xlsx, "--skip-geocode"])
        document = json.loads(out.read_text(encoding="utf-8"))
        assert document["hospitals"], "fixture parsed to zero hospitals"
        for entry in document["hospitals"]:
            assert entry["lat"] is None
            assert entry["lon"] is None
            assert entry["geocode"] == "skipped"
            assert "geocode_accuracy" not in entry

    def test_failed_geocode_stays_null_never_guessed(
        self, fixture_xlsx, tmp_path, monkeypatch
    ):
        import utils.geocoding as geocoding_module

        class _RefusingService:
            def geocode_address(self, _address):
                return None

        monkeypatch.setattr(geocoding_module, "GeocodingService", _RefusingService)
        out = tmp_path / "hospitals.json"
        main(
            [
                "--out",
                str(out),
                "--xlsx",
                fixture_xlsx,
                # Geocoding bills, so the importer asks for a reason.
                "--reason",
                "pytest: CNH import geocoding",
            ]
        )
        document = json.loads(out.read_text(encoding="utf-8"))
        for entry in document["hospitals"]:
            assert entry["lat"] is None
            assert entry["lon"] is None
            assert entry["geocode"] == "failed"

    def test_out_of_region_geocode_is_recorded_failed(
        self, fixture_xlsx, tmp_path, monkeypatch
    ):
        import utils.geocoding as geocoding_module

        class _WrongPlaceService:
            def geocode_address(self, _address):
                # A confident match... in Madrid. Storing it would be a
                # guessed coordinate for a Galician/Asturian hospital.
                return {"lat": 40.4168, "lng": -3.7038, "accuracy": "precise"}

        monkeypatch.setattr(geocoding_module, "GeocodingService", _WrongPlaceService)
        out = tmp_path / "hospitals.json"
        main(
            [
                "--out",
                str(out),
                "--xlsx",
                fixture_xlsx,
                # Geocoding bills, so the importer asks for a reason.
                "--reason",
                "pytest: CNH import geocoding",
            ]
        )
        document = json.loads(out.read_text(encoding="utf-8"))
        for entry in document["hospitals"]:
            assert entry["lat"] is None
            assert entry["lon"] is None
            assert entry["geocode"] == "failed"
            assert "geocode_accuracy" not in entry


def test_normalize_municipality_inverted_articles():
    assert normalize_municipality("Coruña, A") == "A Coruña"
    assert normalize_municipality("Barco de Valdeorras, O") == "O Barco de Valdeorras"
    assert normalize_municipality("Oviedo") == "Oviedo"
